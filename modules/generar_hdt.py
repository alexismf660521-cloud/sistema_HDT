import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from shutil import copyfile

def generar_formularios(archivo, plantilla):

    # =============================
    # ✅ CARGAR DATOS (CORREGIDO)
    # =============================
    df = pd.read_excel(archivo, dtype=str, keep_default_na=False)
    df = df.fillna("").astype(str)

    # =============================
    # ✅ NORMALIZAR COLUMNAS
    # =============================
    df.columns = df.columns.str.strip().str.lower()

    # =============================
    # ✅ MAPEO
    # =============================
    mapeo_estatico = {
        "F2": "nº caso.1",
        "F4": "lugar de origen",
        "F7": "persona asociada",
        "T4": "actual custodia",
        "B44": "notas para el informe pericial",
        "B43": "contexto del caso",
        "B51": "comunicaciones con el cliente",
        "D57": "consecutivo",
        "N57": "folios"
    }

    # =============================
    # ✅ MAPEO DE MUESTRAS
    # =============================
    tipos_muestra_filas = {
        "sangre líquida": 17,
        "orina": 18,
        "lavado vesical": 19,
        "humor vítreo": 20,
        "bilis": 21,
        "contenido gástrico": 22,
        "hematoma": 23,
        "riñón": 24,
        "hígado": 25,
        "frotis": 26,
        "otro": 27,
        "líquidos": 28,
        "sólido": 29
    }

    columnas_muestras = {
        "B": "tipo de emp",
        "D": "emp",
        "I": "recipiente",
        "O": "cantidad",
        "R": "alcoholemia",
        "V": "emp",
        "Z": "psicofármacos-plaguicidas"
    }

    # =============================
    # ✅ CREAR CARPETA
    # =============================
    carpeta = "salida"
    os.makedirs(carpeta, exist_ok=True)

    # =============================
    # 🔁 PROCESAR POR CASO
    # =============================
    for caso_id, grupo in df.groupby("caso_numero"):

        nombre = f"{carpeta}/Formulario_{caso_id}.xlsx"
        copyfile(plantilla, nombre)

        wb = load_workbook(nombre)
        ws = wb.active

        primer = grupo.iloc[0]

        # =============================
        # ✅ CAMPOS ESTÁTICOS
        # =============================
        for celda, campo in mapeo_estatico.items():
            valor = str(primer.get(campo, "")).strip()
            if valor:
                ws[celda] = valor

        # =============================
        # ✅ FECHA (CORREGIDO)
        # =============================
        fecha_str = str(primer.get("fecha de recepcion", "")).strip()

        if fecha_str:
            try:
                fecha = pd.to_datetime(fecha_str, errors="coerce")

                if pd.notna(fecha):
                    ws["T2"] = fecha.year
                    ws["X2"] = f"{fecha.month:02d}"
                    ws["Z2"] = f"{fecha.day:02d}"

            except:
                pass

        # =============================
        # ✅ EMPAQUES
        # =============================
        if "id empaque" in grupo.columns:
            empaques = grupo["id empaque"].dropna().unique()
            empaques = [str(e).strip() for e in empaques if str(e).strip()]
            ws["T7"] = ", ".join(empaques)

        # =============================
        # ✅ MUESTRAS
        # =============================
        for _, fila in grupo.iterrows():

            tipo = str(fila.get("tipo de emp", "")).strip().lower()
            fila_excel = tipos_muestra_filas.get(tipo)

            if fila_excel:
                for col, campo in columnas_muestras.items():
                    valor = str(fila.get(campo, "")).strip()
                    if valor:
                        ws[f"{col}{fila_excel}"] = valor

        # =============================
        # ✅ EVALUACIÓN FIJA
        # =============================
        evaluacion = {
            'B12': "Si X",
            'I12': "Si X",
            'M12': "Si X",
            'S12': "NO X",
            'AA12': "NO X"
        }

        for celda, valor in evaluacion.items():
            ws[celda] = valor

        wb.save(nombre)

    # =============================
    # ✅ ZIP FINAL (CORREGIDO)
    # =============================
    zip_path = "formularios.zip"

    if os.path.exists(zip_path):
        os.remove(zip_path)

    shutil.make_archive("formularios", 'zip', carpeta)

    return zip_path